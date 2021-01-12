# -*- coding:utf-8 -*-
import requests
import json
import chardet
import  locale
import sys,os
import argparse,json
import datetime
import shutil
from openpyxl import load_workbook


def getIds():
    #wb = load_workbook("~/dd/wk/doc/info/前端界面组-成员.xlsx")
    #wb = load_workbook("/drives/d/wk/doc/info/前端界面组-成员.xlsx")
    print(datetime.datetime.now())
    wb = load_workbook("D:\wk\doc\info\前端界面组-成员.xlsx")
    my_sheet = wb.active
    print(my_sheet.title)
    print(my_sheet["F6"].value)
    print(my_sheet.max_row,my_sheet.max_column)
    print(datetime.datetime.now())
    res=[]
    for l in range(3,60):
        nm = my_sheet["F"+str(l)].value
        hwid = my_sheet["O"+str(l)].value
        if nm=="预留" or nm==None:
            continue
        elif nm=="END":
            break
        res.append({"nm":nm,"hwid":hwid})
        print(nm,hwid)
    print(res)
    return res
ef file_names(file_dir):
    L=[]
    for root, dirs, files in os.walk(file_dir):
        for file in files:
            if os.path.splitext(file)[1] == '.xlsx':
                L.append(os.path.join(root, file))
    return L

class Jobs():
    def __init__(self):
        self.js_column_nm={"任务":"B","进度":"C","优先级":"D","风险":"E"}
        self.js_dst_wb_nm = "D:\share\write-share\jobmerge\jobs.xlsx"
        shutil.copyfile("D:\share\write-share\jobmerge\jobs-templete.xlsx",self.js_dst_wb_nm)
        self.js_dst_wb=load_workbook(self.js_dst_wb_nm)
        self.js_dst_wb_sht = self.js_dst_wb.active
        self.js_dst_wb_start_row = 0
        self.initDstStartRow()
        print("initDstStartRow max_row=",self.js_dst_wb_sht.max_row,"startRow=",self.js_dst_wb_start_row)
        print(file_names("D:\share\write-share\jobs"))

    def initDstStartRow(self):
        self.js_dst_wb_start_row = self.js_dst_wb_sht.max_row
        for r in range(1,self.js_dst_wb_sht.max_row+1):
            cellrv = self.js_dst_wb_sht["A"+str(r)].value
            print("initDstStartRow ","A"+str(r),cellrv)
            if cellrv==None or cellrv=="" :
                self.js_dst_wb_start_row = r
                break
    def clearMerge(self):
        pass
    def mergeJob(self,exename="D:\share\write-share\jobs\王崔.xlsx"):
        wb = load_workbook(exename)
        sht = wb.active
        name = os.path.basename(exename).split(".")[0]
        #print(name,sht.title)
        #print(sht.max_row,sht.max_column)
        column_nms={}
        for i in range(sht.max_column+1):
            cell = chr(65+i)+str(1)
            cellv = sht[cell].value
            if cellv == None or cellv == "" :
                continue
            if not cellv in list(self.js_column_nm.keys()) :
                continue
            #print(cell+"="+cellv)
            column_nms[cellv]=chr(65+i)
        print("====next all cell for",exename,"start=",self.js_dst_wb_start_row)
        dst_row = self.js_dst_wb_start_row
        for r in range(2,sht.max_row+1):
            write=False
            for cn in self.js_column_nm.keys():
                cell = column_nms[cn] +str(r)
                #print("cell=",cell)
                cellv = sht[cell].value
                if cellv==None :
                    break
                write=True
                dst_cell = self.js_column_nm[cn]+str(dst_row)
                print(cell,cellv,"==>",dst_cell)
                self.js_dst_wb_sht[dst_cell]=cellv
                self.js_dst_wb_sht["A"+str(dst_row)]=name  #excel name to the first column of dest excel
            if write :
                dst_row+=1
        self.js_dst_wb_start_row = dst_row
        self.js_dst_wb.save(self.js_dst_wb_nm)
        #print("save to ",self.js_dst_wb_nm)
        #wb.save(exename)
    def mergeAllJob(self):
        allExcels = file_names("D:\share\write-share\jobs")
        for xls in allExcels :
            self.mergeJob(xls)

class MonthOutput(Jobs):
    def mergeAllJob(self):
        allExcels = file_names("D:\share\write-share\jobs")
        for xls in allExcels :
            self.mergeJob(xls)

class ArBugs():
    def merge(self,arExcel,bugExcel):
        arExcel="D:\share\prs"

class weeklyReport(Jobs):
    def __init__(self):
        self.js_column_nm={"本周进展":"B","问题、风险":"C"}
        self.js_dst_wb_nm = "D:\share\write-share\周报\merge\week-report.xlsx"
        shutil.copyfile("D:\share\write-share\周报\merge\week-report-templete.xlsx",self.js_dst_wb_nm)
        self.js_dst_wb=load_workbook(self.js_dst_wb_nm)
        self.js_dst_wb_sht = self.js_dst_wb.active
        self.js_dst_wb_start_row = 0
        self.initDstStartRow()
        print("initDstStartRow max_row=",self.js_dst_wb_sht.max_row,"startRow=",self.js_dst_wb_start_row)
    def mergeAllJob(self,week):
        self.js_dst_wb_nm = "D:\share\write-share\周报\merge\week-report-%s%s.xlsx" %("2021",week)
        allExcels = file_names("D:\share\write-share\周报\%s%s" %("2021",week))
        menbers = getIds()
        ln=[ i['nm'] for i in menbers ]
        nown=[]
        #print("mergeAllJob allExcels=",allExcels)
        for xls in allExcels :
            #print("weeklyReport self.mergeJob"+xls)
            nown.append(os.path.basename(xls).split(".")[0])
            self.mergeJob(xls)
        non = list(set(ln)-set(nown))
        print("the follow people not submit report : %s" %(non))

def mainf(args):
    if args.func == "gi":
        getIds()
    elif args.func == "mj":
        j = Jobs()
        j.mergeAllJob()
    elif args.func == "wr":
        j = weeklyReport()
        j.mergeAllJob(args.name)
    elif args.func == "abm":
        ab = ArBugs()
        ab.merge("","")
    else :
        print("unkown fun")

if __name__ == "__main__" :
    parser = argparse.ArgumentParser("process data arrays with numpy and matlib")
    parser.add_argument('-f', "--func", choices=["gi","mj","da","abm","wr"],
                        default="",dest="func",
                        help="qmr:query mr umr:update mr  uc:update cookie wr:weekly report")
    parser.add_argument('-n', action="store",type=str, default="", dest="name",
                       help="name")
    parser.add_argument('--zrid', action="store",type=str, default="0",
                       help="order id")
    parser.add_argument('-g', action="store",type=str, default="", dest="gerriturl",
                       help="gerrit url")
    parser.add_argument('-l', action="store",type=str, default="", dest="gerritline",
                       help="gerrit code change line")

    args = parser.parse_args()
    print("=====need export proxy=====")
    mainf(args)
